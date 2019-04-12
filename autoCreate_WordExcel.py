import sys
import os
import shutil
import time
import openpyxl
import io
import win32com.client
import logging
import coloredlogs


# sys format
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')
# log config
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger("Word")
coloredlogs.install(level='DEBUG')


replaceSourceList = []
sourceColumn = 2
replaceDestList = []
DestColumn = 3
workBook = openpyxl.load_workbook(".\\replace_rule.xlsx")
replaceRule = workBook["replace_rule"]
cutTime = time.strftime('_%Y%m%d-%H_%M_%S', time.localtime(time.time()))
sourceDir = ".\\source_template"


# create replace rule
def create_replace_rule():
    for i in range(3, 102):
        if replaceRule.cell(row=i, column=sourceColumn).value is not None:
            replaceSourceList.append((str(replaceRule.cell(row=i, column=sourceColumn).value)).rstrip(' ').rstrip('\n').rstrip('\r'))
        if replaceRule.cell(row=i, column=DestColumn).value is not None:
            replaceDestList.append((str(replaceRule.cell(row=i, column=DestColumn).value)).rstrip(' ').rstrip('\n').rstrip('\r'))
    logger.info("=== Process Start ===")
    logger.info("Replace source : " + "   ".join(replaceSourceList))
    logger.info("Replace destination : " + "   ".join(replaceDestList))
    global destDir
    destDir = ".\\dest_document\\[autoGen] " + replaceDestList[0] + cutTime
    if (len(replaceSourceList) == len(replaceDestList)) is not True:
        logger.error("!!!!!! [Error] Failed to generate file. Please check file replace_rule.xlsx !!!!!! \n")
        return False
    else:
        return True


# Gen a new copy
def gen_new_copy():
    shutil.copytree(sourceDir, destDir)


def list_all_files(rootdir):
    _files = []
    list = os.listdir(rootdir)
    for i in range(0, len(list)):
        path = os.path.join(rootdir, list[i])
        if os.path.isdir(path):
            _files.extend(list_all_files(path))
        if os.path.isfile(path):
            _files.append(path)
    return _files


def check_file_format(temp_file):
    if 'doc' in (temp_file.lstrip(destDir)).split('.')[-1]:
        return 'doc'
    if 'docx' in (temp_file.lstrip(destDir)).split('.')[-1]:
        return 'doc'
    elif 'wps' in (temp_file.lstrip(destDir)).split('.')[-1]:
        return 'wps'
    elif 'xls' in (temp_file.lstrip(destDir)).split('.')[-1]:
        return 'xls'
    elif 'ignore' in (temp_file.lstrip(destDir)).split('.')[-1]:
        return 'ignore'
    else:
        return 'no_found'


# List all documents and replace according to the rules
def precess_files():
    allFilesList = list_all_files(destDir)
    # enter word environment, maybe don't change env. TODO: FIXME
    word_temp = win32com.client.gencache.EnsureDispatch('word.application')
    i = 0
    for each_file in allFilesList:
        i = i + 1
        # print(each_file)
        # print(destDir)
        # print(each_file.lstrip(destDir))
        logger.info("Start [ " + str(i) + " ] processing : " + each_file.lstrip(destDir))
        if check_file_format(each_file) == 'doc':
            abs_path = os.path.abspath(each_file)
            process_docx(word_temp, abs_path)
            logger.debug("    Replace [ " + str(i) + " ] completed")
        elif check_file_format(each_file) == 'xslx':
            logger.warning("!!!! No susport EXCEL format, please manually modify : " + each_file.lstrip(destDir))
        elif check_file_format(each_file) == 'xls':
            logger.warning("!!!! No susport EXCEL format, please manually modify : " + each_file.lstrip(destDir))
        elif check_file_format(each_file) == 'wps':
            logger.warning("!!!! No susport WPS format, please manually modify : " + each_file.lstrip(destDir))
            # logger.info("Start [ " + str(i) + " ] processing : " + each_file.lstrip(destDir))
            # wps_path = os.path.abspath(each_file)
            # process_wps(wps_path)
            # logger.debug("    Replace [ " + str(i) + " ] completed")
        elif check_file_format(each_file) == 'no_found':
            logger.warning("!!!! No found format, please manually modify : " + each_file.lstrip(destDir))
        else:
            logger.error("!!!! Error format, please manually modify : " + each_file.lstrip(destDir))
    # exit word environment
    word_temp.Quit()
    return


# Docx process lib start
def process_docx(w, file_doc):
    # start word process
#    try:
#        w = win32com.client.gencache.EnsureDispatch('kwps.application')
#        time.sleep(3)
#    except:
#        w = win32com.client.gencache.EnsureDispatch('wps.application')
#        time.sleep(3)
#    else:
    time.sleep(0.1)
#    w = win32com.client.gencache.EnsureDispatch('word.application')
    # running in the background, display program interface, no warning
#    w.Visible = 0
#    w.DisplayAlerts = 0
    # open doc
    # debug
    doc = w.Documents.Open(file_doc)
    for i in range(len(replaceDestList)):
        # replace text body
        w.Selection.Find.ClearFormatting()
        w.Selection.Find.Replacement.ClearFormatting()
        # w.Selection.Find.Execute(replaceSourceList[i], False, False, False, False, False, True, 1, True, replaceDestList[i], 2)
        w.Selection.Find.Execute(replaceSourceList[i], False, False, False, False, False, True, 1, False, replaceDestList[i], 2)
    doc.SaveAs(file_doc)
    w.Documents.Close()
#    w.Quit()
    return
# Docx process lib end


# Docx process lib start
def process_wps(file_doc):
    # start word process
    #try:
    #    w = win32com.client.gencache.EnsureDispatch('kwps.application')
    #except ValueError as e:
    #    w = win32com.client.gencache.EnsureDispatch('wps.application')
    #else:
    w = win32com.client.gencache.EnsureDispatch('wps.application')
    time.sleep(1)
#    w.Visible = 0
#    w.DisplayAlerts = 0
    # open doc
    # debug
    doc = w.Documents.Open(file_doc)
    for i in range(len(replaceDestList)):
        # replace text body
        w.Selection.Find.ClearFormatting()
        w.Selection.Find.Replacement.ClearFormatting()
        # w.Selection.Find.Execute(replaceSourceList[i], False, False, False, False, False, True, 1, True, replaceDestList[i], 2)
        w.Selection.Find.Execute(replaceSourceList[i], False, False, False, False, False, True, 1, False, replaceDestList[i], 2)
    doc.SaveAs(file_doc)
    w.Documents.Close()

    w.Quit()
    return
# Docx process lib end


# Excel process lib start
def process_xlsx(file_xslx):
    # start xlsx process
    ex=win32com.client.Dispatch('Excel.Application')
    # running in the background, # running in the background, display program interface, no warning 
    ex.Visible = 0
    ex.DisplayAlerts = 0
    # open xlsx
    # debug
    # print(file_xslx)
    wk=ex.Workbooks.Open(file_xslx, "utf8")
    for sheet_i in range(wk.sheetnames):
        ws=wk.Worksheets(sheet_i)
        ws.Activate
        for i in range(len(replaceDestList)):
            ex.Selection.Replace(replaceSourceList[i], False, False, False, False, False, True, 1, False, replaceDestList[i], 2)
    wk.Save()
    ex.quit()
    return
# Excel process lib end


if __name__ == "__main__":
    try:
        if create_replace_rule() is True:
            gen_new_copy()
            precess_files()
    except ValueError as e:
        logger.info("=== Process Start ===")
        os.system('pause')
    except ZeroDivisionError as e:
        logger.info("=== Process Start ===")
        os.system('pause')
    else:
        logger.info("=== Process Start ===")
        os.system('pause')
